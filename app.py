from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import importlib.util
import requests
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


TARGET_HEADERS: List[str] = [
    "Розничная старая",
    "Розничная новая",
    "№",
    "Фото",
    "Код",
    "Артикул",
    "Наименование товаров",
    "Остаток",
    "Корп.",
    "Цена опт.",
    "Закупочная",
    "скидка",
    "Акционная цена",
    "наценка",
]
FINAL_HEADERS: List[str] = TARGET_HEADERS + ["Менеджер", "Категория"]

KEYWORDS_FILE = Path("keywords")
KEYWORDS_EXTENSIONS = ("", ".xlsx", ".csv", ".txt")


def setup_logger() -> Tuple[logging.Logger, List[str]]:
    """Настраивает логирование с сохранением сообщений для вывода в UI."""
    logger = logging.getLogger("xlsx_merge")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    messages: List[str] = []

    class ListHandler(logging.Handler):
        def emit(self, record: logging.LogRecord) -> None:
            messages.append(self.format(record))

    formatter = logging.Formatter("%(levelname)s: %(message)s")

    list_handler = ListHandler()
    list_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    logger.addHandler(list_handler)
    logger.addHandler(stream_handler)

    return logger, messages


def normalize_header(value: object) -> str:
    """Нормализует заголовки для сопоставления (нижний регистр, чистка пробелов)."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\.:]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def is_empty(value: object) -> bool:
    """Проверяет, является ли значение пустым."""
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def select_sheet(workbook: Workbook) -> object:
    """Выбирает лист с максимальным числом заполненных строк."""
    best_sheet = workbook.active
    best_count = -1
    for sheet in workbook.worksheets:
        filled = 0
        for row in sheet.iter_rows(values_only=True):
            if any(not is_empty(cell) for cell in row):
                filled += 1
        if filled > best_count:
            best_count = filled
            best_sheet = sheet
    return best_sheet


def read_xlsx_to_rows(
    file_bytes: bytes,
    logger: logging.Logger,
) -> Tuple[List[List[object]], Dict[int, List[bytes]]]:
    """Читает .xlsx в список строк и извлекает изображения."""
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        logger.error("Не удалось открыть .xlsx файл: %s", exc)
        return [], {}

    sheet = select_sheet(workbook)
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    images_by_row: Dict[int, List[bytes]] = {}
    for image in getattr(sheet, "_images", []):
        try:
            anchor = image.anchor
            row_idx = anchor._from.row
            image_bytes = image._data()
            images_by_row.setdefault(row_idx, []).append(image_bytes)
        except Exception as exc:
            logger.warning("Не удалось извлечь изображение из .xlsx: %s", exc)
    return rows, images_by_row


def detect_table_header_row(rows: Sequence[Sequence[object]]) -> Optional[int]:
    """Ищет строку шапки таблицы по ключевым заголовкам."""
    key_number = {"№", "n", "no", "номер"}
    for idx, row in enumerate(rows):
        normalized = {normalize_header(cell) for cell in row if not is_empty(cell)}
        if not normalized:
            continue
        has_number = any(cell in key_number for cell in normalized)
        has_name = "наименование товаров" in normalized
        has_code = "код" in normalized
        if has_number and (has_name or has_code):
            return idx
    return None


def extract_document_header(
    rows: Sequence[Sequence[object]], header_row_idx: int
) -> List[List[object]]:
    """Возвращает строки шапки документа (до строки шапки таблицы)."""
    return [list(row) for row in rows[:header_row_idx]]


def build_source_header_map(header_row: Sequence[object]) -> Dict[str, int]:
    """Создает маппинг нормализованных заголовков на индексы колонок."""
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key and key not in mapping:
            mapping[key] = idx
        if "скидка" in key and "скидка" not in mapping:
            mapping["скидка"] = idx
        if "акц" in key and "акционная цена" not in mapping:
            mapping["акционная цена"] = idx
        if "наценка" in key and "наценка" not in mapping:
            mapping["наценка"] = idx
    return mapping


def find_number_column(header_map: Dict[str, int]) -> Optional[int]:
    """Ищет колонку с номером для фильтрации строк."""
    for key in ["№", "n", "no", "номер"]:
        if key in header_map:
            return header_map[key]
    return None


def map_row_to_target(
    row: Sequence[object],
    source_header_map: Dict[str, int],
    manager_name: str,
    category: str,
) -> List[object]:
    """Преобразует строку исходного файла в строку целевого формата."""
    result: List[object] = []
    result.append(row[0] if len(row) > 0 else "")
    result.append(row[1] if len(row) > 1 else "")

    for header in TARGET_HEADERS[2:]:
        normalized = normalize_header(header)
        source_idx = source_header_map.get(normalized)
        if source_idx is None or source_idx >= len(row):
            result.append("")
        else:
            result.append(row[source_idx])

    result.append(manager_name)
    result.append(category)
    return result


def is_url(value: object) -> bool:
    """Проверяет, что значение похоже на URL."""
    if not value:
        return False
    text = str(value).strip().lower()
    return text.startswith("http://") or text.startswith("https://")


def fetch_image_bytes(url: str, logger: logging.Logger) -> Optional[bytes]:
    """Пытается скачать изображение по URL."""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.content
    except requests.RequestException as exc:
        logger.warning("Не удалось скачать изображение %s: %s", url, exc)
        return None


def write_xlsx(
    document_header: Sequence[Sequence[object]],
    rows: Sequence[Sequence[object]],
    images_for_rows: Sequence[List[bytes]],
    logger: logging.Logger,
) -> Optional[io.BytesIO]:
    """Записывает итоговые данные в .xlsx для дополнительной выгрузки."""
    if importlib.util.find_spec("openpyxl") is None:
        logger.warning("openpyxl не установлен, .xlsx файл не будет создан.")
        return None
    if importlib.util.find_spec("PIL") is None:
        logger.warning("pillow не установлен, изображения в .xlsx не будут вставлены.")

    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Общий"

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=False)
    thin_side = Side(style="thin", color="000000")
    white_side = Side(style="thin", color="FFFFFF")
    thin_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    white_border = Border(
        left=white_side,
        right=white_side,
        top=white_side,
        bottom=white_side,
    )

    manager_col_index = FINAL_HEADERS.index("Менеджер") + 1
    current_row = 1
    for header_row in document_header:
        for col_idx, value in enumerate(header_row, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=value)
            cell.border = white_border
        current_row += 1

    for col_idx, header in enumerate(FINAL_HEADERS, start=1):
        cell = sheet.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if col_idx in (1, 2, manager_col_index):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = header_alignment
        cell.border = thin_border
    current_row += 1

    photo_col_index = FINAL_HEADERS.index("Фото") + 1
    code_col_index = FINAL_HEADERS.index("Код") + 1
    name_col_index = FINAL_HEADERS.index("Наименование товаров") + 1
    discount_col_index = FINAL_HEADERS.index("скидка") + 1
    action_price_col_index = FINAL_HEADERS.index("Акционная цена") + 1
    markup_col_index = FINAL_HEADERS.index("наценка") + 1
    purchase_col_index = FINAL_HEADERS.index("Закупочная") + 1
    category_col_index = FINAL_HEADERS.index("Категория") + 1
    code_max_len = max(len("Код"), 1)
    header_based_widths = {
        purchase_col_index: len("Закупочная") + 2,
        discount_col_index: len("скидка") + 2,
        action_price_col_index: len("Акционная цена") + 2,
        markup_col_index: len("наценка") + 2,
        manager_col_index: len("Менеджер") + 2,
        category_col_index: len("Категория") + 2,
    }

    photo_size_px = 100
    photo_padding_ratio = 1.05
    photo_col_width = (photo_size_px / 7) * photo_padding_ratio
    photo_row_height = (photo_size_px * 0.75) * photo_padding_ratio

    def fit_image_to_cell(image: Image.Image) -> Image.Image:
        """Подгоняет изображение под 100x100, чтобы оно не выходило за пределы ячейки."""
        return image.resize((photo_size_px, photo_size_px))

    def ensure_photo_cell_size(row_idx: int) -> None:
        """Подгоняет размеры ячейки под размер изображения + 5%."""
        column_letter = get_column_letter(photo_col_index)
        sheet.column_dimensions[column_letter].width = max(
            sheet.column_dimensions[column_letter].width or 0,
            photo_col_width,
        )
        sheet.row_dimensions[row_idx].height = max(
            sheet.row_dimensions[row_idx].height or 0,
            photo_row_height,
        )

    for row_index, row in enumerate(rows):
        sheet.row_dimensions[current_row].height = max(
            sheet.row_dimensions[current_row].height or 0,
            15 * 1.05,
        )
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == name_col_index:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif col_idx == manager_col_index:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = body_alignment

            if col_idx == 2 and value is not None:
                try:
                    cell.value = round(float(value))
                except (TypeError, ValueError):
                    pass

            if col_idx == code_col_index and value is not None:
                code_max_len = max(code_max_len, len(str(value)))

            if col_idx == discount_col_index and value is not None:
                try:
                    numeric_value = float(value)
                except (TypeError, ValueError):
                    pass
                else:
                    if numeric_value > 1:
                        numeric_value = numeric_value / 100
                    cell.value = numeric_value
                    cell.number_format = "0%"

            if col_idx == markup_col_index:
                cell.number_format = "0.00"
        images = images_for_rows[row_index] if row_index < len(images_for_rows) else []
        for image_bytes in images:
            if importlib.util.find_spec("PIL") is None:
                logger.warning("pillow не установлен, изображения в .xlsx не будут вставлены.")
                break
            try:
                with Image.open(io.BytesIO(image_bytes)) as img:
                    buffer = io.BytesIO()
                    img = fit_image_to_cell(img)
                    img.save(buffer, format="PNG")
                    buffer.seek(0)
                    openpyxl_image = OpenpyxlImage(buffer)
                    openpyxl_image.anchor = f"{get_column_letter(photo_col_index)}{current_row}"
                    openpyxl_image.width = photo_size_px
                    openpyxl_image.height = photo_size_px
                    ensure_photo_cell_size(current_row)
                    sheet.add_image(openpyxl_image)
            except Exception as exc:
                logger.warning("Не удалось обработать изображение: %s", exc)
        photo_value = row[photo_col_index - 1] if len(row) >= photo_col_index else ""
        if (
            not images
            and is_url(photo_value)
            and importlib.util.find_spec("PIL") is not None
        ):
            image_bytes = fetch_image_bytes(str(photo_value), logger)
            if image_bytes:
                try:
                    with Image.open(io.BytesIO(image_bytes)) as img:
                        buffer = io.BytesIO()
                        img = fit_image_to_cell(img)
                        img.save(buffer, format="PNG")
                        buffer.seek(0)
                        openpyxl_image = OpenpyxlImage(buffer)
                        openpyxl_image.anchor = f"{get_column_letter(photo_col_index)}{current_row}"
                        openpyxl_image.width = photo_size_px
                        openpyxl_image.height = photo_size_px
                        ensure_photo_cell_size(current_row)
                        sheet.add_image(openpyxl_image)
                except Exception as exc:
                    logger.warning("Не удалось обработать изображение: %s", exc)
        current_row += 1

    sheet.column_dimensions[get_column_letter(name_col_index)].width = 60
    sheet.column_dimensions[get_column_letter(code_col_index)].width = max(
        sheet.column_dimensions[get_column_letter(code_col_index)].width or 0,
        code_max_len + 2,
    )
    sheet.column_dimensions[get_column_letter(1)].width = 15
    sheet.column_dimensions[get_column_letter(2)].width = 15
    for col_index, width in header_based_widths.items():
        sheet.column_dimensions[get_column_letter(col_index)].width = max(
            sheet.column_dimensions[get_column_letter(col_index)].width or 0,
            width,
        )

    workbook.save(output)
    output.seek(0)
    return output


@st.cache_data(show_spinner=False)
def load_keywords_table(file_path: Path) -> List[Tuple[str, str]]:
    """Читает таблицу ключевых слов из файла keywords."""
    if not file_path.exists():
        return []

    if file_path.suffix.lower() == ".xlsx":
        try:
            workbook = load_workbook(file_path)
        except Exception:
            return []
        sheet = select_sheet(workbook)
        keywords: List[Tuple[str, str]] = []
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            keyword = str(row[0]).strip() if len(row) > 0 else ""
            category = str(row[1]).strip() if len(row) > 1 else ""
            if keyword:
                keywords.append((keyword.lower(), category))
        return keywords

    content = file_path.read_text(encoding="utf-8").splitlines()
    rows = [line for line in content if line.strip()]
    if not rows:
        return []

    sample = "\n".join(rows[:5])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";\t,")
    except csv.Error:
        dialect = csv.excel

    keywords: List[Tuple[str, str]] = []
    for row in csv.reader(rows, dialect):
        if len(row) < 2:
            continue
        keyword = row[0].strip()
        category = row[1].strip()
        if keyword:
            keywords.append((keyword.lower(), category))
    return keywords


def resolve_keywords_path() -> Optional[Path]:
    """Ищет файл keywords с поддержкой популярных расширений."""
    for ext in KEYWORDS_EXTENSIONS:
        candidate = KEYWORDS_FILE.with_suffix(ext) if ext else KEYWORDS_FILE
        if candidate.exists():
            return candidate
    return None


def get_category(
    product_name: str,
    keywords: Sequence[Tuple[str, str]],
    logger: logging.Logger,
) -> str:
    """Определяет категорию по ключевым словам из файла keywords."""
    if not product_name:
        return ""
    name_lower = product_name.lower()
    for keyword, category in keywords:
        if keyword and keyword in name_lower:
            logger.info("Категория определена по ключевому слову '%s'.", keyword)
            return category
    return ""


def prepare_source_header_map(header_row: Sequence[object]) -> Dict[str, int]:
    """Создает маппинг заголовков исходного файла."""
    return build_source_header_map(header_row)


def main() -> None:
    st.set_page_config(page_title="Объединение .xlsx", layout="wide")
    st.title("Объединение .xlsx файлов")
    st.write(
        "Загрузите несколько файлов Excel (.xlsx), чтобы получить единый файл "
        "с унифицированными колонками."
    )

    logger, log_messages = setup_logger()

    uploaded_files = st.file_uploader(
        "Загрузите .xlsx файлы",
        type=["xlsx"],
        accept_multiple_files=True,
    )
    if uploaded_files:
        with st.spinner("Идёт загрузка файлов..."):
            st.progress(1.0)
    build_xlsx = True

    st.sidebar.header("Загруженные файлы")
    if uploaded_files:
        for uploaded in uploaded_files:
            st.sidebar.write(f"• {uploaded.name}")
    else:
        st.sidebar.write("Файлы не выбраны.")

    if st.button("Объединить"):
        if not uploaded_files:
            st.warning("Не выбраны файлы.")
            return

        keywords_path = resolve_keywords_path()
        keywords = load_keywords_table(keywords_path) if keywords_path else []
        if not keywords:
            logger.warning(
                "Файл keywords не найден или пустой, колонка 'Категория' будет пустой."
            )
        logger.info(
            "Встроенные изображения в .xlsx переносятся, если они присутствуют на листе."
        )

        all_rows: List[List[object]] = []
        all_images: List[List[bytes]] = []
        document_header: List[List[object]] = []
        total_files = len(uploaded_files)
        total_rows_added = 0
        categories_found = 0
        prepared_files: List[Dict[str, object]] = []

        for uploaded in uploaded_files:
            filename = uploaded.name
            logger.info("Файл для обработки: %s", filename)
            try:
                rows, images_by_row = read_xlsx_to_rows(
                    uploaded.getvalue(),
                    logger,
                )
            except Exception as exc:
                logger.error("Не удалось прочитать файл %s: %s", filename, exc)
                continue

            header_row_idx = detect_table_header_row(rows)
            if header_row_idx is None:
                logger.warning("Шапка таблицы не найдена в файле %s", filename)
                continue

            header_row = rows[header_row_idx]
            source_header_map = prepare_source_header_map(header_row)
            logger.info("Найдена строка шапки таблицы: %s", header_row_idx + 1)
            logger.info(
                "Колонки в источнике (нормализованные): %s",
                ", ".join(sorted(source_header_map.keys())),
            )
            if "акционная цена" not in source_header_map:
                logger.warning(
                    "Колонка с ключом 'акц' не найдена в шапке файла %s, "
                    "значения для 'Акционная цена' будут пустыми.",
                    filename,
                )
            number_col_idx = find_number_column(source_header_map)
            if number_col_idx is None:
                logger.warning(
                    "Не удалось определить колонку '№' в файле %s, файл пропущен.",
                    filename,
                )
                continue
            logger.info("Колонка '№' найдена: %s", number_col_idx + 1)

            target_to_source = {}
            missing_targets = []
            for header in TARGET_HEADERS[2:]:
                key = normalize_header(header)
                if key in source_header_map:
                    target_to_source[header] = source_header_map[key] + 1
                else:
                    missing_targets.append(header)
            logger.info("Сопоставление колонок: %s", target_to_source)
            if missing_targets:
                logger.warning("Колонки без источника: %s", ", ".join(missing_targets))

            data_rows = rows[header_row_idx + 1 :]
            prepared_files.append(
                {
                    "filename": filename,
                    "rows": rows,
                    "header_row_idx": header_row_idx,
                    "source_header_map": source_header_map,
                    "number_col_idx": number_col_idx,
                    "images_by_row": images_by_row,
                }
            )

        for file_index, file_data in enumerate(prepared_files, start=1):
            filename = file_data["filename"]
            rows = file_data["rows"]
            header_row_idx = file_data["header_row_idx"]
            source_header_map = file_data["source_header_map"]
            number_col_idx = file_data["number_col_idx"]
            manager_name = str(filename).rsplit(".", 1)[0]
            images_by_row = file_data["images_by_row"]
            logger.info("Обработка файла: %s", filename)

            if not document_header:
                document_header = extract_document_header(rows, header_row_idx)

            data_rows = rows[header_row_idx + 1 :]
            for row_index, row in enumerate(data_rows, start=1):
                if row_index <= 3:
                    logger.info(
                        "Пример строки %s: код=%s, наименование=%s, скидка=%s, закупочная=%s",
                        row_index,
                        row[source_header_map.get("код", -1)]
                        if "код" in source_header_map
                        and source_header_map["код"] < len(row)
                        else "",
                        row[source_header_map.get("наименование товаров", -1)]
                        if "наименование товаров" in source_header_map
                        and source_header_map["наименование товаров"] < len(row)
                        else "",
                        row[source_header_map.get("скидка", -1)]
                        if "скидка" in source_header_map
                        and source_header_map["скидка"] < len(row)
                        else "",
                        row[source_header_map.get("закупочная", -1)]
                        if "закупочная" in source_header_map
                        and source_header_map["закупочная"] < len(row)
                        else "",
                    )
                number_value = row[number_col_idx] if number_col_idx < len(row) else ""
                if is_empty(number_value):
                    continue

                first_col_value = row[0] if len(row) > 0 else ""
                if is_empty(first_col_value):
                    continue

                product_name = (
                    row[source_header_map.get("наименование товаров", -1)]
                    if "наименование товаров" in source_header_map
                    and source_header_map["наименование товаров"] < len(row)
                    else ""
                )

                category = get_category(str(product_name), keywords, logger)
                if category:
                    categories_found += 1

                mapped_row = map_row_to_target(
                    row=row,
                    source_header_map=source_header_map,
                    manager_name=manager_name,
                    category=category,
                )
                all_rows.append(mapped_row)
                image_row_index = header_row_idx + row_index
                all_images.append(images_by_row.get(image_row_index, []))
                total_rows_added += 1


        if not all_rows:
            logger.warning("Нет данных для сохранения.")
            st.text_area("Логи", "\n".join(log_messages), height=200)
            return

        output_xlsx = None
        if build_xlsx:
            output_xlsx = write_xlsx(document_header, all_rows, all_images, logger)

        st.success("Объединение завершено.")
        st.write(
            f"Файлов обработано: {total_files}. "
            f"Строк собрано: {total_rows_added}. "
            f"Категорий найдено: {categories_found}."
        )

        if output_xlsx is not None:
            st.download_button(
                label="Скачать Акция ОБЩИЙ.xlsx",
                data=output_xlsx,
                file_name="Акция ОБЩИЙ.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.warning("Не удалось сформировать .xlsx файл для скачивания.")

        st.text_area("Логи", "\n".join(log_messages), height=200)


if __name__ == "__main__":
    main()
